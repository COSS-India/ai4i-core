#!/usr/bin/env python3
"""
Fires 500 OpenAI-compatible chat-completions requests at the AI4I dev
gateway (dev.ai4inclusion.org), split across 4 hardcoded (tenant API key,
model ID) combinations — see COMBINATIONS below — and records
input/output/token counts for every call.

Every request sends the EXACT SAME input string, so prompt_tokens should
stay identical across all runs (barring server-side tokenizer nondeterminism).
completion_tokens/output text can vary per call since model sampling is not
necessarily deterministic.

Setup
-----
    pip install requests openpyxl

Then run:

    python scripts/llm-load-test.py

Requests run concurrently across a small thread pool (see CONCURRENCY below)
since each call is I/O-bound. Results are printed to the console as they
complete (order is completion order, not request order — see the "req #N"
tag on each line) and written incrementally (row-by-row, flushed after every
request) to OUTPUT_CSV so a Ctrl+C partway through still leaves a usable file.

On top of the CSV, every run also APPENDS a summary block (start/end time,
per-combo request counts, avg latency, avg prompt tokens) to OUTPUT_XLSX —
run this script as many times as you like and each run adds a fresh block
below the previous ones, so the sheet builds up a history you can scroll
through and copy start/end timestamps out of into Grafana.
"""

import csv
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone

import requests

# ─────────────────────────────────────────────────────────────────────────────
# Config — edit these before running
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://dev.ai4inclusion.org"
CHAT_COMPLETIONS_PATH = "/api/v1/chat/completions"

# 4 hardcoded (tenant API key, model ID) combinations. Each combo gets a
# fixed slice of the 500 total requests (125 each below) — adjust
# num_requests per combo if you want an uneven split, just keep the total
# at 500 (asserted below).
COMBINATIONS = [
    {
        "label": "combo1: v4 / key-a7f2",
        "model_id": "aug-20-llm-2/Automation-ServiceID-v4/",
        "api_key": "9ee97751721f78c0975efe6d63b0a7f2",
        "num_requests": 125,
    },
    {
        "label": "combo2: v5 / key-fc8ce",
        "model_id": "aug-20-llm-2/Automation-ServiceID-v5/",
        "api_key": "553599bad2181f9ed4c3e8eef64fc8ce",
        "num_requests": 125,
    },
    {
        "label": "combo3: v6 / key-a710d29",
        "model_id": "aug-20-llm-2/Automation-ServiceID-v6/",
        "api_key": "039993c453d471ba1318162a1a710d29",
        "num_requests": 125,
    },
    {
        "label": "combo4: v4 / key-fc8ce",
        "model_id": "aug-20-llm-2/Automation-ServiceID-v4/",
        "api_key": "553599bad2181f9ed4c3e8eef64fc8ce",
        "num_requests": 125,
    },
]

# Same input every time -> same prompt token count every time.
INPUT_TEXT = "Summarize the importance of clean drinking water for rural communities briefly."
assert 50 <= len(INPUT_TEXT) <= 100, f"INPUT_TEXT must be 50-100 chars, got {len(INPUT_TEXT)}"

NUM_REQUESTS = sum(c["num_requests"] for c in COMBINATIONS)
assert NUM_REQUESTS == 500, f"COMBINATIONS must add up to 500 requests, got {NUM_REQUESTS}"

# How many requests run at once, across all combos combined. HTTP calls are
# I/O-bound, so a thread pool is enough (no need for asyncio). Keep this
# modest on a shared dev environment — raise it if the endpoint tolerates
# more, lower it if you start seeing timeouts/429s/connection errors.
CONCURRENCY = 10
REQUEST_TIMEOUT_S = 60
OUTPUT_CSV = os.path.join(os.path.dirname(__file__), "llm-load-test-results.csv")
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), "llm-load-test-runs.xlsx")

# ─────────────────────────────────────────────────────────────────────────────

CSV_FIELDS = [
    "index",
    "combo",
    "model_id",
    "timestamp_utc",
    "status",
    "http_status",
    "input",
    "output",
    "prompt_tokens",
    "completion_tokens",
    "total_tokens",
    "latency_ms",
    "error",
]

XLSX_SUMMARY_HEADERS = [
    "Combo",
    "Model ID",
    "API Key",
    "Requests",
    "Success",
    "Failures",
]

TIME_FMT = "%Y-%m-%d %H:%M:%S"
# Run start/end are reported in IST (not UTC) in both the console output and
# the xlsx summary, since that's what gets pasted into Grafana's time range
# picker. Per-request rows in OUTPUT_CSV stay UTC (timestamp_utc column) —
# only the whole-run start/end times shown to the user are converted.
IST = timezone(timedelta(hours=5, minutes=30), name="IST")


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds")


def fmt_ist(dt: datetime) -> str:
    return dt.astimezone(IST).strftime(TIME_FMT)


def mask_key(key: str) -> str:
    """Shorten a secret for display in the shared xlsx log (e.g. printed/screen-shared)."""
    if len(key) <= 10:
        return key
    return f"{key[:6]}...{key[-4:]}"


def send_one(session: requests.Session, index: int, combo: dict) -> dict:
    """Send a single chat-completions request (for the given combo) and return a result row."""
    payload = {
        "model": combo["model_id"],
        "messages": [{"role": "user", "content": INPUT_TEXT}],
        "stream": False,
    }
    headers = {
        "Authorization": f"Bearer {combo['api_key']}",
        "Content-Type": "application/json",
    }

    row = {
        "index": index,
        "combo": combo["label"],
        "model_id": combo["model_id"],
        "timestamp_utc": now_utc_iso(),
        "status": "error",
        "http_status": "",
        "input": INPUT_TEXT,
        "output": "",
        "prompt_tokens": "",
        "completion_tokens": "",
        "total_tokens": "",
        "latency_ms": "",
        "error": "",
    }

    start = time.perf_counter()
    try:
        resp = session.post(
            f"{BASE_URL}{CHAT_COMPLETIONS_PATH}",
            json=payload,
            headers=headers,
            timeout=REQUEST_TIMEOUT_S,
        )
        elapsed_ms = (time.perf_counter() - start) * 1000
        row["latency_ms"] = round(elapsed_ms, 1)
        row["http_status"] = resp.status_code

        if resp.status_code >= 400:
            row["error"] = resp.text[:500]
            return row

        data = resp.json()
        output_text = (
            data.get("choices", [{}])[0].get("message", {}).get("content", "")
        )
        usage = data.get("usage", {}) or {}

        row["status"] = "ok"
        row["output"] = output_text
        row["prompt_tokens"] = usage.get("prompt_tokens", "")
        row["completion_tokens"] = usage.get("completion_tokens", "")
        row["total_tokens"] = usage.get("total_tokens", "")
        return row

    except requests.RequestException as exc:
        row["latency_ms"] = round((time.perf_counter() - start) * 1000, 1)
        row["error"] = str(exc)
        return row


def append_xlsx_summary(run_start: datetime, run_end: datetime, rows: list) -> None:
    """
    Append one summary block for this run to OUTPUT_XLSX — creating the
    workbook (with a title row) on first use, otherwise loading the existing
    one and writing the new block a couple of rows below whatever is already
    there, so every run's block stacks under the previous ones.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        next_row = ws.max_row + 2  # blank separator row before the new block
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "LLM load test runs"
        next_row = 1

    bold = Font(bold=True)

    # Whole-run totals — avg latency and avg prompt tokens are reported once
    # for the entire run (all combos combined), not broken out per combo.
    total_requests = sum(r["num_requests"] for r in rows)
    total_success = sum(r["success"] for r in rows)
    total_failures = sum(r["failures"] for r in rows)
    total_latency_sum = sum(r["latency_sum"] for r in rows)
    total_latency_n = sum(r["latency_n"] for r in rows)
    total_prompt_tokens_sum = sum(r["prompt_tokens_sum"] for r in rows)
    total_prompt_tokens_n = sum(r["prompt_tokens_n"] for r in rows)
    overall_avg_latency = (total_latency_sum / total_latency_n) if total_latency_n else ""
    overall_avg_prompt_tokens = (total_prompt_tokens_sum / total_prompt_tokens_n) if total_prompt_tokens_n else ""

    # ── Block title + start/end time range (copy-paste straight into Grafana) ──
    ws.cell(row=next_row, column=1, value=f"Run — {fmt_ist(run_start)} IST").font = bold
    next_row += 1
    ws.cell(row=next_row, column=1, value="Start time (IST)")
    ws.cell(row=next_row, column=2, value=fmt_ist(run_start))
    ws.cell(row=next_row, column=3, value="End time (IST)")
    ws.cell(row=next_row, column=4, value=fmt_ist(run_end))
    next_row += 1
    ws.cell(row=next_row, column=1, value="Total requests")
    ws.cell(row=next_row, column=2, value=total_requests)
    ws.cell(row=next_row, column=3, value="Avg latency (ms)")
    ws.cell(row=next_row, column=4, value=round(overall_avg_latency, 1) if overall_avg_latency != "" else "")
    ws.cell(row=next_row, column=5, value="Avg prompt tokens")
    ws.cell(row=next_row, column=6, value=round(overall_avg_prompt_tokens, 1) if overall_avg_prompt_tokens != "" else "")
    next_row += 2

    # ── Per-combo table — request/success/failure counts only; avg latency
    # and avg prompt tokens are whole-run figures above, not per combo. ──
    for col, header in enumerate(XLSX_SUMMARY_HEADERS, start=1):
        ws.cell(row=next_row, column=col, value=header).font = bold
    header_row = next_row
    next_row += 1

    for r in rows:
        ws.cell(row=next_row, column=1, value=r["label"])
        ws.cell(row=next_row, column=2, value=r["model_id"])
        ws.cell(row=next_row, column=3, value=mask_key(r["api_key"]))
        ws.cell(row=next_row, column=4, value=r["num_requests"])
        ws.cell(row=next_row, column=5, value=r["success"])
        ws.cell(row=next_row, column=6, value=r["failures"])
        next_row += 1

    total_font = Font(bold=True)
    ws.cell(row=next_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=next_row, column=4, value=total_requests).font = total_font
    ws.cell(row=next_row, column=5, value=total_success).font = total_font
    ws.cell(row=next_row, column=6, value=total_failures).font = total_font

    # Widen columns a bit on first creation so the sheet is readable out of the box.
    if header_row == 5:  # only true on the very first block written to a fresh workbook
        widths = [26, 40, 20, 10, 9, 9]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    wb.save(OUTPUT_XLSX)


def main() -> None:
    run_start = datetime.now(timezone.utc)
    print("=" * 78)
    print("AI4I LLM load test")
    print(f"  endpoint       : {BASE_URL}{CHAT_COMPLETIONS_PATH}")
    print(f"  requests       : {NUM_REQUESTS} (concurrency={CONCURRENCY})")
    print(f"  input ({len(INPUT_TEXT)} chars): {INPUT_TEXT!r}")
    for c in COMBINATIONS:
        print(f"  combo          : {c['label']:<22} model={c['model_id']:<38} requests={c['num_requests']}")
    print(f"  results csv    : {OUTPUT_CSV}")
    print(f"  runs xlsx      : {OUTPUT_XLSX}")
    print(f"  start time     : {fmt_ist(run_start)} IST")
    print("=" * 78)

    # Flat (index, combo) work list — 125 rows for combo1, then 125 for
    # combo2, etc., but all 500 are submitted to the same thread pool so
    # combos run concurrently, not one after another.
    work_items = []
    per_combo_index = 1
    for combo in COMBINATIONS:
        for _ in range(combo["num_requests"]):
            work_items.append((per_combo_index, combo))
            per_combo_index += 1

    # Per-combo running stats, keyed by combo label, for the xlsx summary.
    combo_stats = {
        c["label"]: {
            **c,
            "success": 0,
            "failures": 0,
            "latency_sum": 0.0,
            "latency_n": 0,
            "prompt_tokens_sum": 0,
            "prompt_tokens_n": 0,
        }
        for c in COMBINATIONS
    }

    ok_count = 0
    err_count = 0
    completed = 0
    latencies = []
    # Requests complete out of order under concurrency; this lock keeps CSV
    # rows, combo stats and console lines from interleaving across worker threads.
    io_lock = threading.Lock()

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f, requests.Session() as session:
        # Default adapter pool caps at 10 connections; size it to CONCURRENCY
        # so raising CONCURRENCY doesn't trigger "Connection pool is full"
        # warnings (requests would still work, just with extra contention).
        adapter = requests.adapters.HTTPAdapter(pool_connections=CONCURRENCY, pool_maxsize=CONCURRENCY)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        f.flush()

        with ThreadPoolExecutor(max_workers=CONCURRENCY) as pool:
            futures = [pool.submit(send_one, session, idx, combo) for idx, combo in work_items]

            for future in as_completed(futures):
                row = future.result()
                stats = combo_stats[row["combo"]]

                with io_lock:
                    completed += 1
                    writer.writerow(row)
                    f.flush()

                    if row["status"] == "ok":
                        ok_count += 1
                        stats["success"] += 1
                        if row["prompt_tokens"] != "":
                            stats["prompt_tokens_sum"] += row["prompt_tokens"]
                            stats["prompt_tokens_n"] += 1
                        stats["latency_sum"] += row["latency_ms"]
                        stats["latency_n"] += 1
                        latencies.append(row["latency_ms"])

                        print(
                            f"[{completed:>4}/{NUM_REQUESTS}] OK  (req #{row['index']:>3} {row['combo']})  "
                            f"latency={row['latency_ms']:>7.1f}ms  "
                            f"prompt_tokens={row['prompt_tokens']:<5} "
                            f"completion_tokens={row['completion_tokens']:<5} "
                            f"total_tokens={row['total_tokens']:<5}"
                        )
                    else:
                        err_count += 1
                        stats["failures"] += 1
                        print(
                            f"[{completed:>4}/{NUM_REQUESTS}] FAIL (req #{row['index']:>3} {row['combo']})  "
                            f"http_status={row['http_status']} error={row['error']}"
                        )

    run_end = datetime.now(timezone.utc)
    duration_s = (run_end - run_start).total_seconds()

    print("=" * 78)
    print("Run complete")
    print(f"  start time (IST) : {fmt_ist(run_start)}")
    print(f"  end time (IST)   : {fmt_ist(run_end)}")
    print(f"  total duration   : {duration_s:.1f}s")
    print(f"  successes        : {ok_count}/{NUM_REQUESTS}")
    print(f"  failures         : {err_count}/{NUM_REQUESTS}")
    for label, stats in combo_stats.items():
        print(
            f"    {label:<22} requests={stats['num_requests']:<4} "
            f"success={stats['success']:<4} failures={stats['failures']:<4}"
        )
    # Avg latency / avg prompt tokens are whole-run figures (all combos
    # combined), not broken out per combo — matches the xlsx summary.
    if latencies:
        print(f"  avg latency      : {sum(latencies) / len(latencies):.1f}ms  (whole run)")
        print(f"  min/max latency  : {min(latencies):.1f}ms / {max(latencies):.1f}ms")
    total_prompt_tokens_n = sum(s["prompt_tokens_n"] for s in combo_stats.values())
    if total_prompt_tokens_n:
        total_prompt_tokens_sum = sum(s["prompt_tokens_sum"] for s in combo_stats.values())
        print(f"  avg prompt tokens: {total_prompt_tokens_sum / total_prompt_tokens_n:.1f}  (whole run)")
    print(f"  results csv      : {OUTPUT_CSV}")

    try:
        append_xlsx_summary(run_start, run_end, list(combo_stats.values()))
        print(f"  runs xlsx        : {OUTPUT_XLSX}  (new block appended)")
    except ImportError:
        print("  runs xlsx        : SKIPPED — install with `pip install openpyxl` to enable")
    except PermissionError:
        print(
            f"  runs xlsx        : FAILED — {OUTPUT_XLSX} is locked (likely still open in "
            "Excel). Close it and rerun; the CSV above was still written."
        )
    except Exception as exc:
        # Anything else (corrupt/half-written xlsx from a previous crash, disk
        # full, etc.) must not look like a silent no-op — surface it loudly
        # instead of letting it propagate past the final "=" banner unseen.
        print(f"  runs xlsx        : FAILED — {type(exc).__name__}: {exc}")
    print("=" * 78)


if __name__ == "__main__":
    main()
